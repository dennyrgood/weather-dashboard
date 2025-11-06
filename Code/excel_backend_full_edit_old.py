from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import shutil
import logging

app = Flask(__name__)
CORS(app)

# Logging setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration
EXCEL_FILE = os.path.expanduser("D:/OneDrive/MS/MoviesShows.xlsx")

# Mapping of HTML field names to Excel column indices (1-indexed)
INPUT_MAP = {
    'code': 1, 'title': 6, 'col_g': 7, 'col_h': 8, 'col_i': 9,
    'col_j': 10, 'col_k': 11, 'col_l': 12, 'col_m': 13
}
EDITABLE_COLUMNS = list(INPUT_MAP.values())
COLUMN_LETTERS = {v: k for k, v in INPUT_MAP.items()}

# ============ CORE UTILITY FUNCTIONS ============

def create_backup(original_file):
    """Create timestamped backup before modifications."""
    if not os.path.exists(original_file):
        return False, f"File not found: {original_file}"
    
    base, ext = os.path.splitext(original_file)
    timestamp = datetime.now().strftime(" - Backup %Y %m %d %H %M %S")
    backup_file = f"{base}{timestamp}{ext}"
    
    try:
        shutil.copy2(original_file, backup_file)
        return True, backup_file
    except Exception as e:
        return False, f"Backup failed: {str(e)}"

def copy_cell_properties(source_cell, target_cell, copy_number_format=True):
    """Copy formatting from source to target cell."""
    if source_cell.has_style:
        target_cell.style = source_cell.style
    if copy_number_format and source_cell.number_format:
        target_cell.number_format = source_cell.number_format

def copy_formulas(ws, source_row, target_row):
    """Copy formulas from B,C,D,E with updated row references."""
    formula_cols = [2, 3, 4, 5]
    old_row_str = str(source_row)
    new_row_str = str(target_row)
    columns_to_update = ['A', 'B']
    replacement_targets = []
    
    for col in columns_to_update:
        replacement_targets.append((f"{col}{old_row_str}", f"{col}{new_row_str}"))
        replacement_targets.append((f"{col.lower()}{old_row_str}", f"{col.lower()}{new_row_str}"))
    
    for col_idx in formula_cols:
        source_cell = ws.cell(row=source_row, column=col_idx)
        target_cell = ws.cell(row=target_row, column=col_idx)
        
        if source_cell.data_type == 'f' and source_cell.value:
            formula_text = source_cell.value
            updated_formula = formula_text
            for old_ref, new_ref in replacement_targets:
                updated_formula = updated_formula.replace(old_ref, new_ref)
            target_cell.value = updated_formula
        else:
            target_cell.value = source_cell.value

def parse_coordinate(coord_str):
    """Parse Excel coordinate (e.g., 'A10') into (col_letter, row_num)."""
    col = ""
    row_str = ""
    for char in coord_str:
        if char.isalpha():
            col += char.upper()
        elif char.isdigit():
            row_str += char
    if not col or not row_str:
        raise ValueError(f"Invalid coordinate: {coord_str}")
    return (col, int(row_str))

def parse_date(date_str):
    """Parse flexible date formats. Returns date object or None."""
    if not date_str:
        return None
    
    formats = [
        '%d-%b-%Y', '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d',
        '%d/%m/%Y', '%d/%m/%y'
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    return None

# ============ CORE EXCEL CRUD FUNCTIONS ============

def get_table_info(ws):
    """Utility to get table range and name if present."""
    if ws.tables:
        table_name = list(ws.tables.keys())[0]
        table = ws.tables[table_name]
        return table_name, table.ref, table.tableStyleInfo if hasattr(table, 'tableStyleInfo') else None
    return None, None, None

def get_data_from_excel(target_file):
    """Reads all data from the active worksheet and returns it as a list of dicts."""
    data = []
    try:
        wb = load_workbook(target_file, data_only=True)
        ws = wb.active
        
        table_name, table_range, _ = get_table_info(ws)
        
        if table_range:
            start_col, start_row = parse_coordinate(table_range.split(':')[0])
            end_col, end_row = parse_coordinate(table_range.split(':')[1])
            
            # Start reading from the row *after* the header row
            for row_idx in range(start_row + 1, end_row + 1):
                row_data = {"row_index": row_idx}
                
                for col_idx in range(1, 14):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    
                    # Format dates consistently: return as dd-mmm-yyyy string for frontend
                    if col_idx == 10 and isinstance(cell_value, date):
                        cell_value = cell_value.strftime('%d-%b-%Y') if cell_value else None
                    elif isinstance(cell_value, datetime):
                        cell_value = cell_value.isoformat()
                    
                    field_name = COLUMN_LETTERS.get(col_idx, get_column_letter(col_idx))
                    row_data[field_name] = cell_value
                
                data.append(row_data)
        
        return True, data
    except Exception as e:
        logger.error(f"Excel read error: {str(e)}")
        return False, f"Error reading data: {str(e)}"

def update_row_in_excel(target_file, row_index, row_data):
    """Updates the content of a specific row in the Excel sheet."""
    try:
        wb = load_workbook(target_file, data_only=False)
        ws = wb.active
        
        # Validate row exists
        if row_index < 1 or row_index > ws.max_row:
            return False, f"Row {row_index} is out of range (max: {ws.max_row})"
        
        # Validate required fields
        code_str = row_data.get('code', '')
        title = row_data.get('title', '')
        
        if not code_str or not title:
            return False, "Code and Title are required"
        
        # Validate code is numeric
        try:
            float(code_str)
        except (ValueError, TypeError):
            return False, "Code must be a valid number"
        
        # Update cells
        for key, col_idx in INPUT_MAP.items():
            if key in row_data:
                value = row_data[key]
                target_cell = ws.cell(row=row_index, column=col_idx)
                
                if col_idx == 10:  # Col J (Date)
                    if value:
                        parsed_date = parse_date(str(value))
                        if not parsed_date:
                            return False, f"Invalid date format: {value}"
                        target_cell.value = parsed_date
                        target_cell.number_format = 'dd-mmm-yyyy'
                    else:
                        target_cell.value = None
                elif col_idx == 1:  # Col A (Code)
                    try:
                        target_cell.value = float(value)
                    except (ValueError, TypeError):
                        target_cell.value = value
                else:
                    target_cell.value = value
        
        # Refresh formulas in B,C,D,E by copying them to same row (triggers recalculation)
        copy_formulas(ws, row_index, row_index)
        
        wb.save(target_file)
        return True, f"Row {row_index} updated successfully"
    except Exception as e:
        logger.error(f"Excel update error: {str(e)}")
        return False, f"Error updating row: {str(e)}"

def delete_row_from_excel(target_file, row_index):
    """Deletes a specific row and updates the table reference."""
    try:
        wb = load_workbook(target_file, data_only=False)
        ws = wb.active
        
        # Validate row exists
        if row_index < 1 or row_index > ws.max_row:
            return False, f"Row {row_index} is out of range (max: {ws.max_row})"
        
        table_name, table_range, table_style = get_table_info(ws)
        
        if table_range:
            start_coord_str, end_coord_str = table_range.split(':')
            start_col_letter, row_start_num = parse_coordinate(start_coord_str)
            end_col_letter, max_row = parse_coordinate(end_coord_str)
            
            # Check if the row to delete is within the table data range (excluding header)
            if row_index > row_start_num and row_index <= max_row:
                ws.delete_rows(row_index)
                
                # Update table range
                if max_row - row_start_num > 1:  # More than just header remains
                    new_max_row = max_row - 1
                    new_ref = f"{start_col_letter}{row_start_num}:{end_col_letter}{new_max_row}"
                    del ws.tables[table_name]
                    new_table = Table(displayName=table_name, ref=new_ref)
                    if table_style:
                        new_table.tableStyleInfo = table_style
                    ws.add_table(new_table)
                else:
                    # All data rows deleted, remove table object
                    del ws.tables[table_name]
                
                wb.save(target_file)
                return True, f"Row {row_index} deleted successfully"
            else:
                return False, f"Row {row_index} is outside the data range or is a header row"
        
        ws.delete_rows(row_index)
        wb.save(target_file)
        return True, f"Row {row_index} deleted successfully"

    except Exception as e:
        logger.error(f"Excel delete error: {str(e)}")
        return False, f"Error deleting row: {str(e)}"

def insert_row_to_excel(target_file, new_row_data):
    """Add new row to Excel with formatting and formula preservation."""
    try:
        wb = load_workbook(target_file, data_only=False)
        ws = wb.active
        
        table_name, table_range, table_style = get_table_info(ws)
        max_row = ws.max_row
        row_start_num = 1
        
        if table_range:
            start_coord_str, end_coord_str = table_range.split(':')
            start_col_letter, row_start_num = parse_coordinate(start_coord_str)
            end_col_letter, max_row = parse_coordinate(end_coord_str)
        else:
            start_col_letter, end_col_letter = 'A', get_column_letter(max(EDITABLE_COLUMNS))
        
        last_existing_row = max_row
        new_row_number = last_existing_row + 1
        formula_source_row = last_existing_row if last_existing_row > row_start_num else row_start_num
        
        # Copy properties and formulas from source row
        for col_idx in range(1, 14):
            source_cell = ws.cell(row=formula_source_row, column=col_idx)
            target_cell = ws.cell(row=new_row_number, column=col_idx)
            is_column_j = (col_idx == 10)
            copy_cell_properties(source_cell, target_cell, copy_number_format=not is_column_j)
        
        if formula_source_row > row_start_num:
            copy_formulas(ws, formula_source_row, new_row_number)
        
        # Set new data values
        data_to_cols = [
            (new_row_data[0], INPUT_MAP['code']),
            (new_row_data[1], INPUT_MAP['title']),
            (new_row_data[2], INPUT_MAP['col_g']),
            (new_row_data[3], INPUT_MAP['col_h']),
            (new_row_data[4], INPUT_MAP['col_i']),
            (new_row_data[5], INPUT_MAP['col_j']),
            (new_row_data[6], INPUT_MAP['col_k']),
            (new_row_data[7], INPUT_MAP['col_l']),
            (new_row_data[8], INPUT_MAP['col_m'])
        ]
        
        for value, col_idx in data_to_cols:
            target_cell = ws.cell(row=new_row_number, column=col_idx, value=value)
            if col_idx == 10 and isinstance(value, date):
                target_cell.number_format = 'dd-mmm-yyyy'
        
        # Update table reference
        if table_name and table_range:
            new_ref = f"{start_col_letter}{row_start_num}:{end_col_letter}{new_row_number}"
            del ws.tables[table_name]
            new_table = Table(displayName=table_name, ref=new_ref)
            if table_style:
                new_table.tableStyleInfo = table_style
            ws.add_table(new_table)
        
        wb.save(target_file)
        return True, f"Row {new_row_number} added successfully"
    
    except Exception as e:
        logger.error(f"Excel processing error: {str(e)}")
        return False, f"Error: {str(e)}"

# ============ FLASK ROUTES ============

@app.route('/api/data', methods=['GET'])
def get_data():
    """Fetches all data from the active sheet for display/editing."""
    success, result = get_data_from_excel(EXCEL_FILE)
    if success:
        return jsonify({"status": "success", "data": result}), 200
    else:
        return jsonify({"status": "error", "message": result}), 500

@app.route('/api/add', methods=['POST'])
def add_data():
    """Handles form submission to add a new row."""
    try:
        data = request.json
        code_str = data.get('code', '').strip()
        title = data.get('title', '').strip()
        
        if not code_str or not title:
            return jsonify({"status": "error", "message": "Code and Title are required"}), 400
        
        try:
            code = float(code_str)
        except ValueError:
            return jsonify({"status": "error", "message": "Code must be a number"}), 400
        
        col_g = data.get('col_g', '').strip()
        col_h = data.get('col_h', '').strip()
        col_i = data.get('col_i', 'Download').strip()
        col_j_str = data.get('col_j', '').strip()
        col_k = data.get('col_k', '').strip()
        col_l = data.get('col_l', '').strip()
        col_m = data.get('col_m', '').strip()
        
        col_j = parse_date(col_j_str) if col_j_str else None
        if col_j_str and not col_j:
            return jsonify({"status": "error", "message": f"Invalid date format: {col_j_str}"}), 400
        
        # Create backup (warning only, don't block)
        backup_ok, backup_msg = create_backup(EXCEL_FILE)
        if not backup_ok:
            logger.warning(f"Backup warning: {backup_msg}")
        
        new_row_data = [code, title, col_g, col_h, col_i, col_j, col_k, col_l, col_m]
        success, message = insert_row_to_excel(EXCEL_FILE, new_row_data)
        
        if success:
            logger.info(f"Data added: {title}")
            return jsonify({"status": "success", "message": message}), 200
        else:
            return jsonify({"status": "error", "message": message}), 500
    
    except Exception as e:
        logger.error(f"API /add error: {str(e)}")
        return jsonify({"status": "error", "message": f"Server error: {str(e)}"}), 500

@app.route('/api/update', methods=['POST'])
def update_data():
    """Handles updating an existing row."""
    try:
        data = request.json
        row_index = data.get('row_index')
        
        if not row_index:
            return jsonify({"status": "error", "message": "Row index is required for update"}), 400
        
        # Create backup (warning only, don't block)
        backup_ok, backup_msg = create_backup(EXCEL_FILE)
        if not backup_ok:
            logger.warning(f"Backup warning: {backup_msg}")
        
        success, message = update_row_in_excel(EXCEL_FILE, row_index, data)
        
        if success:
            logger.info(f"Row {row_index} updated.")
            return jsonify({"status": "success", "message": message}), 200
        else:
            return jsonify({"status": "error", "message": message}), 400

    except Exception as e:
        logger.error(f"API /update error: {str(e)}")
        return jsonify({"status": "error", "message": f"Server error: {str(e)}"}), 500

@app.route('/api/delete', methods=['POST'])
def delete_data():
    """Handles deleting a row."""
    try:
        data = request.json
        row_index = data.get('row_index')
        
        if not row_index:
            return jsonify({"status": "error", "message": "Row index is required for delete"}), 400
        
        # Create backup (warning only, don't block)
        backup_ok, backup_msg = create_backup(EXCEL_FILE)
        if not backup_ok:
            logger.warning(f"Backup warning: {backup_msg}")

        success, message = delete_row_from_excel(EXCEL_FILE, row_index)
        
        if success:
            logger.info(f"Row {row_index} deleted.")
            return jsonify({"status": "success", "message": message}), 200
        else:
            return jsonify({"status": "error", "message": message}), 400
            
    except Exception as e:
        logger.error(f"API /delete error: {str(e)}")
        return jsonify({"status": "error", "message": f"Server error: {str(e)}"}), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Simple health check endpoint."""
    return jsonify({"status": "ok"}), 200

if __name__ == '__main__':
    try:
        import flask_cors
    except ImportError:
        print("Installing flask-cors...")
        os.system("pip install flask-cors")
    
    logger.info(f"Starting Flask server")
    logger.info(f"Excel file: {EXCEL_FILE}")
    logger.info(f"Listening on 0.0.0.0:5000")
    
    app.run(host='0.0.0.0', port=5000, debug=False)